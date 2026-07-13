from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from tqdm import tqdm


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from spurbreast_repro.config import load_config, project_path  # noqa: E402
from spurbreast_repro.data import SPLITS, SliceRecord, discover_records  # noqa: E402
from spurbreast_repro.utils import write_json  # noqa: E402


EXPECTED_IMAGES = {
    ("training", 0): 4781,
    ("training", 1): 4781,
    ("validation", 0): 1788,
    ("validation", 1): 1788,
    ("test", 0): 3394,
    ("test", 1): 3394,
}
EXPECTED_PATIENTS = {
    ("training", 0): 200,
    ("training", 1): 200,
    ("validation", 0): 75,
    ("validation", 1): 75,
    ("test", 0): 150,
    ("test", 1): 150,
}
FIELD_CODE_MAP = {0: 1.494, 1: 1.5, 2: 2.8936, 3: 3.0, "0": 1.494, "1": 1.5, "2": 2.8936, "3": 3.0}


def metadata_field_map(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Data"]
    mapping: dict[str, float] = {}
    for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=6, values_only=True):
        patient_id, field_code = row[0], row[5]
        if not patient_id:
            continue
        if field_code not in FIELD_CODE_MAP:
            raise ValueError(f"Unexpected field-strength code for {patient_id}: {field_code!r}")
        mapping[str(patient_id)] = FIELD_CODE_MAP[field_code]
    workbook.close()
    return mapping


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_internal_manifest(path: Path, records: list[SliceRecord]) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["relative_path", "split", "label", "patient_id", "slice_id", "field_strength_t"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for record in records:
            writer.writerow({name: getattr(record, name) for name in fields})
    return sha256_file(path)


def audit_records(
    records: list[SliceRecord],
    verify_images: bool,
    hash_images: bool,
) -> dict:
    image_counts = Counter((record.split, record.label) for record in records)
    patient_sets: dict[tuple[str, int], set[str]] = defaultdict(set)
    for record in records:
        patient_sets[(record.split, record.label)].add(record.patient_id)
    failures: list[str] = []
    for key, expected in EXPECTED_IMAGES.items():
        if image_counts[key] != expected:
            failures.append(f"Image count {key}: {image_counts[key]} != {expected}")
    for key, expected in EXPECTED_PATIENTS.items():
        if len(patient_sets[key]) != expected:
            failures.append(f"Patient count {key}: {len(patient_sets[key])} != {expected}")

    split_patients = {
        split: patient_sets[(split, 0)] | patient_sets[(split, 1)] for split in SPLITS
    }
    overlap_counts = {
        "training_validation": len(split_patients["training"] & split_patients["validation"]),
        "training_test": len(split_patients["training"] & split_patients["test"]),
        "validation_test": len(split_patients["validation"] & split_patients["test"]),
    }
    if any(overlap_counts.values()):
        failures.append(f"Cross-split patient overlap: {overlap_counts}")
    within_label_intersections = {
        split: len(patient_sets[(split, 0)] & patient_sets[(split, 1)]) for split in SPLITS
    }
    if within_label_intersections != {"training": 0, "validation": 0, "test": 150}:
        failures.append(f"Unexpected within-split label intersections: {within_label_intersections}")

    field_counts: dict[str, dict[str, int]] = {}
    for split in SPLITS:
        for label in (0, 1):
            key = f"{split}/{label}"
            counter = Counter(
                f"{record.field_strength_t:g}T"
                for record in records
                if record.split == split and record.label == label
            )
            field_counts[key] = dict(sorted(counter.items()))
    expected_fields = {
        "training/0": {"3T": 4781},
        "training/1": {"1.5T": 4781},
        "validation/0": {"3T": 1788},
        "validation/1": {"1.5T": 1788},
    }
    for key, expected in expected_fields.items():
        if field_counts[key] != expected:
            failures.append(f"Unexpected shortcut direction {key}: {field_counts[key]}")

    modes: Counter[str] = Counter()
    sizes: Counter[str] = Counter()
    corrupt_images = 0
    if verify_images:
        for record in tqdm(records, desc="Verifying PNG files"):
            try:
                with Image.open(record.path) as image:
                    modes[image.mode] += 1
                    sizes[f"{image.width}x{image.height}"] += 1
                    image.verify()
            except Exception:
                corrupt_images += 1
        if corrupt_images:
            failures.append(f"Corrupt images: {corrupt_images}")

    digest_to_splits: dict[str, set[str]] = defaultdict(set)
    digest_counts: Counter[str] = Counter()
    if hash_images:
        for record in tqdm(records, desc="Hashing PNG files"):
            digest = sha256_file(record.path)
            digest_counts[digest] += 1
            digest_to_splits[digest].add(record.split)
    duplicate_groups = sum(count > 1 for count in digest_counts.values())
    cross_split_duplicate_groups = sum(len(splits) > 1 for splits in digest_to_splits.values())
    if cross_split_duplicate_groups:
        failures.append(f"Exact image hashes cross splits: {cross_split_duplicate_groups}")

    patient_field_counts = Counter()
    seen_patient_field: set[tuple[str, float]] = set()
    for record in records:
        pair = (record.patient_id, float(record.field_strength_t))
        if pair not in seen_patient_field:
            patient_field_counts[f"{record.split}/{record.field_strength_t:g}T"] += 1
            seen_patient_field.add(pair)
    expected_patient_fields = {
        "training/1.5T": 200,
        "training/3T": 200,
        "validation/1.5T": 75,
        "validation/3T": 75,
        "test/1.5T": 71,
        "test/3T": 79,
    }
    if dict(sorted(patient_field_counts.items())) != expected_patient_fields:
        failures.append(
            "Unexpected patient field-strength distribution: "
            f"{dict(sorted(patient_field_counts.items()))}"
        )

    summary = {
        "status": "passed" if not failures else "failed",
        "failures": failures,
        "total_images": len(records),
        "total_unique_patients": len({record.patient_id for record in records}),
        "image_counts": {f"{split}/{label}": image_counts[(split, label)] for split in SPLITS for label in (0, 1)},
        "patient_counts": {f"{split}/{label}": len(patient_sets[(split, label)]) for split in SPLITS for label in (0, 1)},
        "unique_split_patients": {split: len(split_patients[split]) for split in SPLITS},
        "cross_split_overlap_counts": overlap_counts,
        "within_split_label_intersections": within_label_intersections,
        "field_strength_image_counts": field_counts,
        "field_strength_patient_counts": dict(sorted(patient_field_counts.items())),
        "image_modes": dict(modes),
        "image_sizes": dict(sizes),
        "corrupt_images": corrupt_images,
        "duplicate_hash_groups": duplicate_groups,
        "cross_split_duplicate_hash_groups": cross_split_duplicate_groups,
        "verified_images": verify_images,
        "hashed_images": hash_images,
    }
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Build and audit the official SpurBreast split")
    parser.add_argument("--config", default="configs/reproduction.yaml")
    parser.add_argument("--verify-images", action="store_true")
    parser.add_argument("--hash-images", action="store_true")
    args = parser.parse_args()
    config, _ = load_config(args.config)
    dataset_dir = project_path(config["data"]["dataset_dir"])
    metadata_file = project_path(config["data"]["metadata_file"])
    manifest_file = project_path(config["data"]["manifest_file"])
    field_map = metadata_field_map(metadata_file)
    records = [
        record
        for split in SPLITS
        for record in discover_records(dataset_dir, split, field_map=field_map)
    ]
    missing = sorted({record.patient_id for record in records if record.field_strength_t is None})
    if missing:
        raise RuntimeError(f"Patients missing from TCIA metadata: {missing[:10]}")
    manifest_sha256 = write_internal_manifest(manifest_file, records)
    field_map_path = manifest_file.parent / "field_strength_by_patient.csv"
    field_map_path.parent.mkdir(parents=True, exist_ok=True)
    with field_map_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["patient_id", "field_strength_t"])
        for patient_id in sorted({record.patient_id for record in records}):
            writer.writerow([patient_id, field_map[patient_id]])
    summary = audit_records(records, args.verify_images, args.hash_images)
    summary["manifest_sha256"] = manifest_sha256
    summary["metadata_cohort_patients"] = len(field_map)
    report_path = PROJECT_ROOT / "reports" / "tables" / "data_audit_summary.json"
    write_json(report_path, summary)
    print(json.dumps(summary, indent=2, sort_keys=True))
    if summary["status"] != "passed":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
